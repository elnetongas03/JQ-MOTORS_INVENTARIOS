
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib import colors
from fpdf import FPDF
import os
import sys
import unicodedata
from pathlib import Path
from datetime import datetime
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from reportlab.pdfgen import canvas
# --------------------
# RUTAS ABSOLUTAS
# --------------------
CARPETA_DATOS = Path.home() / "Desktop" / "Archivos"
CARPETA_EXCEL = CARPETA_DATOS / "Excel"
CARPETA_EXPORT = CARPETA_DATOS / "Export"
LOGO_DIR = CARPETA_DATOS / "LOGO"
CARPETA_DESCARGAS = Path.home() / "Downloads"
# Crear carpetas si no existen
for carpeta in [CARPETA_DATOS, CARPETA_EXCEL, CARPETA_EXPORT, LOGO_DIR, CARPETA_DESCARGAS]:
    os.makedirs(carpeta, exist_ok=True)
# --------------------
# ARCHIVOS EXCEL
# --------------------
ARCHIVO_INVENTARIO = CARPETA_EXCEL / "inventario.xlsx"
ARCHIVO_VENTAS = CARPETA_EXCEL / "ventas.xlsx"
ARCHIVO_PEDIDOS = CARPETA_EXCEL / "pedidos.xlsx"
ARCHIVO_TALLER = CARPETA_EXCEL / "taller.xlsx"
ARCHIVO_COTIZACIONES = CARPETA_EXCEL / "cotizaciones.xlsx"
ARCHIVO_MOTOS = CARPETA_EXCEL / "motos_insumos.xlsx"
# --------------------
# Funciones auxiliares
# --------------------
def quitar_acentos(texto):
    if not isinstance(texto, str):
        return texto
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
def _create_empty_excel(path: Path, columns):
    df = pd.DataFrame(columns=columns)
    df.to_excel(path, index=False, engine="openpyxl")
def load_file(path: Path, columns):
    if path.exists():
        try:
            return pd.read_excel(path, engine="openpyxl", dtype=str).fillna("")
        except Exception:
            return pd.DataFrame(columns=columns)
    else:
        _create_empty_excel(path, columns)
        return pd.DataFrame(columns=columns)
def obtener_estado_codigo(codigo, cantidad_total):
    try:
        return int(cantidad_total), 0
    except Exception:
        return 0, 0
def save_df(path: Path, df: pd.DataFrame):
    df.to_excel(path, index=False, engine="openpyxl")
def load_inventario_file():
    return load_file(ARCHIVO_INVENTARIO, ["codigo", "descripcion", "ubicacion", "stock", "precio"])
def load_ventas_file():
    return load_file(ARCHIVO_VENTAS, ["fecha", "forma_pago", "codigo", "cantidad", "p_unitario", "precio", "total"])
def habilitar_copia_treeview(tree):
    def copiar(event):
        seleccion = tree.selection()
        if not seleccion: return
        texto = ""
        for item in seleccion:
            texto += "\t".join([str(tree.set(item, col)) for col in tree["columns"]]) + "\n"
        tree.clipboard_clear()
        tree.clipboard_append(texto)
    tree.bind("<Control-c>", copiar)
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)
# --------------------
# TREEVIEW ESTILO PERSONALIZADO
# --------------------
def estilo_treeview(root, fondo="#1ec2df"):
    style = ttk.Style(root)
    style.theme_use('clam')
    if fondo.lower() in ["#dd1111",'white']:
        fg_color = 'black'
        sel_bg = "#abe01a"
    else:
        fg_color = 'white'
        sel_bg = '#016630'
    style.configure("Treeview",
                    background=fondo,
                    foreground=fg_color,
                    fieldbackground=fondo,
                    rowheight=24,
                    font=("Segoe UI", 10))

    style.configure("Treeview.Heading",
                    background=fondo,
                    foreground=fg_color,
                    font=("Segoe UI", 10, "bold"))
    style.map("Treeview",
              background=[('selected', sel_bg)],
              foreground=[('selected', fg_color)])
# --------------------
# STYLES GENERALES
# --------------------
def aplicar_estilos(app):
    estilo_treeview(root, fondo="#11d8be")  # Fondo blanco por defecto
    style = ttk.Style(root)
    root.configure(bg="#0753c4")  # Fondo general blanco
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd

# Suponiendo que estas funciones y constantes ya existen en tu proyecto
# quitar_acentos, obtener_estado_codigo, load_inventario_file, save_df, ARCHIVO_INVENTARIO

class Stock(ttk.Frame):
    def __init__(self, parent, controller=None):
        super().__init__(parent)
        self.controller = controller
        ttk.Label(self, text='STOCK', font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=6, pady=6)

        # ------------------- BUSCAR POR CÓDIGO -------------------
        frame = ttk.Frame(self)
        frame.pack(fill='x', padx=6, pady=(5,0))
        ttk.Label(frame, text='Código:').grid(row=0, column=0, sticky='w')
        self.entry_codigo = ttk.Entry(frame, width=25)
        self.entry_codigo.grid(row=0, column=1, padx=4)
        ttk.Button(frame, text='Buscar', command=self.buscar_codigo).grid(row=0, column=2, padx=4)

        # ------------------- BUSCAR POR DESCRIPCIÓN -------------------
        frame2 = ttk.Frame(self)
        frame2.pack(fill='x', padx=6, pady=(6,0))
        ttk.Label(frame2, text='Descripción:').grid(row=0, column=0, sticky='w')
        self.entry_desc = ttk.Entry(frame2, width=40)
        self.entry_desc.grid(row=0, column=1, padx=4)
        ttk.Button(frame2, text='Buscar descripción', command=self.buscar_descripcion).grid(row=0, column=2, padx=4)

        # ------------------- BOTONES ARRIBA A LA DERECHA -------------------
        frame_top_btns = ttk.Frame(self)
        frame_top_btns.pack(fill='x', padx=6)
        ttk.Button(frame_top_btns, text='Importar', command=self.importar_inventario).pack(side='right', padx=4)

        # ------------------- AGREGAR / DESCONTAR -------------------
        frame_desc = ttk.LabelFrame(self, text="AGREGAR / DESCONTAR")
        frame_desc.pack(fill='x', padx=6, pady=6)
        ttk.Label(frame_desc, text="Código:").grid(row=0, column=0, padx=4, pady=4, sticky='w')
        self.desc_codigo = ttk.Entry(frame_desc, width=15)
        self.desc_codigo.grid(row=0, column=1, padx=4, pady=4)
        ttk.Label(frame_desc, text="Cantidad:").grid(row=0, column=2, padx=4, pady=4, sticky='w')
        self.desc_cantidad = ttk.Entry(frame_desc, width=10)
        self.desc_cantidad.grid(row=0, column=3, padx=4, pady=4)
        ttk.Button(frame_desc, text="Agregar", command=self.agregar_refaccion).grid(row=0, column=4, padx=4, pady=4)
        ttk.Button(frame_desc, text="Descontar", command=self.descontar_refaccion).grid(row=0, column=5, padx=4, pady=4)

        # ------------------- AGREGAR / BORRAR ARTÍCULO -------------------
        frame_art = ttk.LabelFrame(self, text="Agregar / Borrar Artículo")
        frame_art.pack(fill='x', padx=6, pady=6)
        ttk.Label(frame_art, text="Código:").grid(row=0, column=0, padx=4, pady=2)
        self.art_codigo = ttk.Entry(frame_art, width=12)
        self.art_codigo.grid(row=0, column=1, padx=4, pady=2)
        ttk.Label(frame_art, text="Descripción:").grid(row=0, column=2, padx=4, pady=2)
        self.art_desc = ttk.Entry(frame_art, width=20)
        self.art_desc.grid(row=0, column=3, padx=4, pady=2)
        ttk.Label(frame_art, text="Ubicación:").grid(row=0, column=4, padx=4, pady=2)
        self.art_ubi = ttk.Entry(frame_art, width=12)
        self.art_ubi.grid(row=0, column=5, padx=4, pady=2)
        ttk.Label(frame_art, text="Stock:").grid(row=1, column=0, padx=4, pady=2)
        self.art_stock = ttk.Entry(frame_art, width=12)
        self.art_stock.grid(row=1, column=1, padx=4, pady=2)
        ttk.Label(frame_art, text="Precio:").grid(row=1, column=2, padx=4, pady=2)
        self.art_precio = ttk.Entry(frame_art, width=12)
        self.art_precio.grid(row=1, column=3, padx=4, pady=2)
        ttk.Button(frame_art, text="Agregar/Actualizar", command=self.agregar_articulo).grid(row=1, column=4, padx=4, pady=2)
        ttk.Button(frame_art, text="Borrar Seleccionado", command=self.borrar_seleccionado).grid(row=1, column=5, padx=4, pady=2)

        # ------------------- TREEVIEW PRINCIPAL -------------------
        cols = ["codigo", "descripcion", "ubicacion", "stock", "precio", "libres", "en_taller", "nuevas_entradas"]
        self.tree = ttk.Treeview(self, columns=cols, show='headings', height=14)
        for c in cols:
            self.tree.heading(c, text=c.capitalize())
            self.tree.column(c, width=120, anchor='center')
        self.tree.pack(fill='both', expand=True, padx=6, pady=6)
        self.cargar_datos()

    # --------------------------------------------------------
    # IMPORTAR INVENTARIO
    # --------------------------------------------------------
    def importar_inventario(self):
        archivo = filedialog.askopenfilename(title='Seleccionar Excel', filetypes=[('Excel','*.xlsx;*.xls')])
        if not archivo: return
        try:
            df_new = pd.read_excel(archivo, engine='openpyxl', dtype=str).fillna('')
            if self.controller:
                self.controller.inventario_df = pd.concat(
                    [self.controller.inventario_df, df_new], ignore_index=True
                )
                if 'codigo' in self.controller.inventario_df.columns:
                    self.controller.inventario_df.drop_duplicates(subset=['codigo'], keep='last', inplace=True)
            save_df(ARCHIVO_INVENTARIO, self.controller.inventario_df)
            self.cargar_datos()
            messagebox.showinfo('Éxito','Inventario importado')
        except Exception as e:
            messagebox.showerror('Error', str(e))

    # --------------------------------------------------------
    # EXPORTAR INVENTARIO
    # --------------------------------------------------------
    def exportar_inventario(self):
        archivo = filedialog.asksaveasfilename(title='Guardar Excel', defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')])
        if not archivo: return
        try:
            df = load_inventario_file()
            df.to_excel(archivo, index=False, engine='openpyxl')
            messagebox.showinfo('Éxito','Inventario exportado')
        except Exception as e:
            messagebox.showerror('Error', str(e))

    # --------------------------------------------------------
    # CARGAR DATOS
    # --------------------------------------------------------
    def cargar_datos(self):
        # usar el DataFrame del controlador si existe
        if self.controller and hasattr(self.controller, 'inventario_df'):
            df = self.controller.inventario_df.copy()
        else:
            df = load_inventario_file()

        if df.empty:
            self.tree.delete(*self.tree.get_children())
            return

        for col in ['libres','en_taller','nuevas_entradas']:
            df[col] = df.get(col,0)

        if 'codigo' in df.columns:
            df['codigo_norm'] = df['codigo'].astype(str).apply(lambda x: quitar_acentos(x).strip().upper())
            df.drop_duplicates(subset=['codigo_norm'], keep='last', inplace=True)
            df.drop(columns=['codigo_norm'], inplace=True)

        self.tree.delete(*self.tree.get_children())

        for _, r in df.iterrows():
            try:
                stock = int(r.get("stock",0))
            except:
                stock = 0
            libres, en_taller = obtener_estado_codigo(r.get("codigo",""), stock)
            vals = (
                r.get('codigo',''),
                r.get('descripcion',''),
                r.get('ubicacion',''),
                stock,
                r.get('precio',0),
                libres,
                en_taller,
                r.get('nuevas_entradas',0)
            )
            self.tree.insert('', 'end', values=vals)


    # --------------------------------------------------------
    # BUSCAR POR CÓDIGO
    # --------------------------------------------------------
    def buscar_codigo(self):
        codigo = quitar_acentos(self.entry_codigo.get().strip()).upper()
        if not codigo:
            messagebox.showinfo('Atención','Escribe un código para buscar')
            return
        df = load_inventario_file()
        if 'codigo' not in df.columns: return
        df['codigo_clean'] = df['codigo'].astype(str).apply(lambda x: quitar_acentos(x).upper())
        r = df[df['codigo_clean']==codigo]
        if r.empty:
            messagebox.showinfo('Atención',f'Código {codigo} no encontrado')
            return
        row = r.iloc[0]
        try: stock = int(row.get('stock',0))
        except: stock = 0
        libres, en_taller = obtener_estado_codigo(row.get("codigo",""), stock)
        self.tree.delete(*self.tree.get_children())
        self.tree.insert('', 'end', values=(
            row.get('codigo',''),
            row.get('descripcion',''),
            row.get('ubicacion',''),
            stock,
            row.get('precio',0),
            libres,
            en_taller,
            row.get('nuevas_entradas',0)
        ))
        # rellenar formulario
        self.art_codigo.delete(0,'end'); self.art_codigo.insert(0,row.get('codigo',''))
        self.art_desc.delete(0,'end'); self.art_desc.insert(0,row.get('descripcion',''))
        self.art_ubi.delete(0,'end'); self.art_ubi.insert(0,row.get('ubicacion',''))
        self.art_stock.delete(0,'end'); self.art_stock.insert(0,str(stock))
        self.art_precio.delete(0,'end'); self.art_precio.insert(0,str(row.get('precio',0)))

    # --------------------------------------------------------
    # BUSCAR POR DESCRIPCIÓN
    # --------------------------------------------------------
    def buscar_descripcion(self):
        desc = quitar_acentos(self.entry_desc.get().strip()).upper()
        if not desc:
            messagebox.showinfo('Atención','Escribe descripción')
            return
        df = load_inventario_file()
        if 'descripcion' not in df.columns: return
        df['desc_clean'] = df['descripcion'].astype(str).apply(lambda x: quitar_acentos(x).upper())
        r = df[df['desc_clean'].str.contains(desc, na=False)]
        self.tree.delete(*self.tree.get_children())
        for _, row in r.iterrows():
            try: stock = int(row.get('stock',0))
            except: stock = 0
            libres, en_taller = obtener_estado_codigo(row.get("codigo",""), stock)
            self.tree.insert('', 'end', values=(
                row.get('codigo',''),
                row.get('descripcion',''),
                row.get('ubicacion',''),
                stock,
                row.get('precio',0),
                libres,
                en_taller,
                row.get('nuevas_entradas',0)
            ))

    # --------------------------------------------------------
    # AGREGAR / DESCONTAR REFACCIÓN
    # --------------------------------------------------------
    def agregar_refaccion(self):
        codigo = self.desc_codigo.get().strip().upper()
        try: cantidad = int(self.desc_cantidad.get())
        except: 
            messagebox.showwarning("Atención","Cantidad inválida")
            return
        if not codigo or cantidad <= 0:
            messagebox.showwarning("Atención","Ingrese código y cantidad válida")
            return
        df = load_inventario_file()
        mask = df['codigo'].astype(str).str.upper() == codigo
        if mask.any():
            idx = df[mask].index[0]
            df.at[idx,'stock'] = int(df.at[idx,'stock']) + cantidad
            save_df(ARCHIVO_INVENTARIO, df)
            self.cargar_datos()
            messagebox.showinfo("OK",f"Agregado {cantidad} a {codigo}")
        else:
            messagebox.showwarning("No encontrado","Código no encontrado")

    def descontar_refaccion(self):
        codigo = self.desc_codigo.get().strip().upper()
        try: cantidad = int(self.desc_cantidad.get())
        except:
            messagebox.showwarning("Atención","Cantidad inválida")
            return
        if not codigo or cantidad <= 0:
            messagebox.showwarning("Atención","Ingrese código y cantidad válida")
            return
        df = load_inventario_file()
        mask = df['codigo'].astype(str).str.upper() == codigo
        if mask.any():
            idx = df[mask].index[0]
            df.at[idx,'stock'] = max(int(df.at[idx,'stock']) - cantidad, 0)
            save_df(ARCHIVO_INVENTARIO, df)
            self.cargar_datos()
            messagebox.showinfo("OK",f"Descontado {cantidad} de {codigo}")
        else:
            messagebox.showwarning("No encontrado","Código no encontrado")

    # --------------------------------------------------------
    # AGREGAR / BORRAR ARTÍCULO
    # --------------------------------------------------------
    def agregar_articulo(self):
        codigo = self.art_codigo.get().strip()
        desc = self.art_desc.get().strip()
        ubi = self.art_ubi.get().strip()
        try: stock = int(self.art_stock.get())
        except:
            messagebox.showwarning("Atención","Stock inválido")
            return
        try: precio = float(self.art_precio.get())
        except:
            messagebox.showwarning("Atención","Precio inválido")
            return
        if not codigo:
            messagebox.showwarning("Atención","Código requerido")
            return
        df = load_inventario_file()
        mask = df['codigo'].astype(str).str.upper() == codigo.upper()
        if mask.any():
            idx = df[mask].index[0]
            df.at[idx,'descripcion'] = desc
            df.at[idx,'ubicacion'] = ubi
            df.at[idx,'stock'] = stock
            df.at[idx,'precio'] = precio
        else:
            df.loc[len(df)] = [codigo, desc, ubi, stock, precio]
        save_df(ARCHIVO_INVENTARIO, df)
        self.cargar_datos()
        messagebox.showinfo("OK","Artículo agregado/actualizado")

    def borrar_seleccionado(self):
        sel = self.tree.selection()
        if not sel: return
        df = load_inventario_file()
        for s in sel:
            codigo = self.tree.item(s)['values'][0]
            df = df[df['codigo'].astype(str).str.upper() != codigo.upper()]
        save_df(ARCHIVO_INVENTARIO, df)
        self.cargar_datos()
        messagebox.showinfo("OK","Artículo(s) borrado(s)")

         # --------------------------------------------------------
    # SINCRONIZACIÓN AUTOMÁTICA CON RENDER
    # --------------------------------------------------------
    def enviar_a_render(self, articulo):
        """
        Envía un artículo actualizado al servidor en Render.
        """
        url = "https://jq-motors-inventarios.onrender.com/inventario"
        try:
            resp = requests.post(url, json=articulo, timeout=5)
            if resp.status_code == 200:
                print(f"Sincronizado con Render: {articulo['codigo']}")
            else:
                print(f"Error al enviar a Render: {resp.status_code}")
        except Exception as e:
            print(f"No se pudo enviar a Render: {e}")

    # Modificar las funciones existentes para enviar cambios automáticamente
    def agregar_refaccion(self):
        codigo = self.desc_codigo.get().strip().upper()
        try: cantidad = int(self.desc_cantidad.get())
        except: 
            messagebox.showwarning("Atención","Cantidad inválida")
            return
        if not codigo or cantidad <= 0:
            messagebox.showwarning("Atención","Ingrese código y cantidad válida")
            return
        df = load_inventario_file()
        mask = df['codigo'].astype(str).str.upper() == codigo
        if mask.any():
            idx = df[mask].index[0]
            df.at[idx,'stock'] = int(df.at[idx,'stock']) + cantidad
            save_df(ARCHIVO_INVENTARIO, df)
            self.cargar_datos()
            messagebox.showinfo("OK",f"Agregado {cantidad} a {codigo}")

            # enviar a Render
            row = df.loc[idx]
            self.enviar_a_render({
                "codigo": row['codigo'],
                "descripcion": row.get('descripcion',''),
                "ubicacion": row.get('ubicacion',''),
                "stock": row.get('stock',0),
                "precio": row.get('precio',0)
            })
        else:
            messagebox.showwarning("No encontrado","Código no encontrado")

    def descontar_refaccion(self):
        codigo = self.desc_codigo.get().strip().upper()
        try: cantidad = int(self.desc_cantidad.get())
        except:
            messagebox.showwarning("Atención","Cantidad inválida")
            return
        if not codigo or cantidad <= 0:
            messagebox.showwarning("Atención","Ingrese código y cantidad válida")
            return
        df = load_inventario_file()
        mask = df['codigo'].astype(str).str.upper() == codigo
        if mask.any():
            idx = df[mask].index[0]
            df.at[idx,'stock'] = max(int(df.at[idx,'stock']) - cantidad, 0)
            save_df(ARCHIVO_INVENTARIO, df)
            self.cargar_datos()
            messagebox.showinfo("OK",f"Descontado {cantidad} de {codigo}")

            # enviar a Render
            row = df.loc[idx]
            self.enviar_a_render({
                "codigo": row['codigo'],
                "descripcion": row.get('descripcion',''),
                "ubicacion": row.get('ubicacion',''),
                "stock": row.get('stock',0),
                "precio": row.get('precio',0)
            })
        else:
            messagebox.showwarning("No encontrado","Código no encontrado")

    def agregar_articulo(self):
        codigo = self.art_codigo.get().strip()
        desc = self.art_desc.get().strip()
        ubi = self.art_ubi.get().strip()
        try: stock = int(self.art_stock.get())
        except:
            messagebox.showwarning("Atención","Stock inválido")
            return
        try: precio = float(self.art_precio.get())
        except:
            messagebox.showwarning("Atención","Precio inválido")
            return
        if not codigo:
            messagebox.showwarning("Atención","Código requerido")
            return
        df = load_inventario_file()
        mask = df['codigo'].astype(str).str.upper() == codigo.upper()
        if mask.any():
            idx = df[mask].index[0]
            df.at[idx,'descripcion'] = desc
            df.at[idx,'ubicacion'] = ubi
            df.at[idx,'stock'] = stock
            df.at[idx,'precio'] = precio
        else:
            df.loc[len(df)] = [codigo, desc, ubi, stock, precio]
            idx = df.index[-1]

        save_df(ARCHIVO_INVENTARIO, df)
        self.cargar_datos()
        messagebox.showinfo("OK","Artículo agregado/actualizado")

        # enviar a Render
        row = df.loc[idx]
        self.enviar_a_render({
            "codigo": row['codigo'],
            "descripcion": row.get('descripcion',''),
            "ubicacion": row.get('ubicacion',''),
            "stock": row.get('stock',0),
            "precio": row.get('precio',0)
        })

# --------------------
# Clase Ventas
# --------------------
class Ventas(ttk.Frame):
    def __init__(self, parent, controller=None):
        super().__init__(parent)
        self.controller = controller
        ttk.Label(self, text='VENTAS', font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=6, pady=6)
        # ------------------
        # Panel de ingreso de venta
        # ------------------
        frame_venta = ttk.LabelFrame(self, text="Agregar a la Venta")
        frame_venta.pack(fill='x', padx=6, pady=6)
        ttk.Label(frame_venta, text="Código:").grid(row=0, column=0, padx=4, pady=4)
        self.cod_entry = ttk.Entry(frame_venta, width=15)
        self.cod_entry.grid(row=0, column=1, padx=4, pady=4)
        self.cod_entry.bind("<FocusOut>", self.completar_datos)  # Autocompleta al salir del campo
        # Para autocompletar mientras se escribe, usar:
        # self.cod_entry.bind("<KeyRelease>", self.completar_datos)
        ttk.Label(frame_venta, text="Descripción:").grid(row=0, column=2, padx=4, pady=4)
        self.desc_entry = ttk.Entry(frame_venta, width=30)
        self.desc_entry.grid(row=0, column=3, padx=4, pady=4)
        ttk.Label(frame_venta, text="Precio:").grid(row=0, column=4, padx=4, pady=4)
        self.precio_entry = ttk.Entry(frame_venta, width=10)
        self.precio_entry.grid(row=0, column=5, padx=4, pady=4)
        ttk.Label(frame_venta, text="Cantidad:").grid(row=0, column=6, padx=4, pady=4)
        self.cant_entry = ttk.Entry(frame_venta, width=8)
        self.cant_entry.grid(row=0, column=7, padx=4, pady=4)
        ttk.Button(frame_venta, text="Agregar a Venta", command=self.agregar_a_venta).grid(row=0, column=8, padx=4, pady=4)
        # -------------------------
        # Forma de pago
        # -------------------------
        frame_pago = ttk.LabelFrame(self, text="Forma de pago", padding=10)
        frame_pago.pack(fill="x", padx=10, pady=10)
        self.forma_pago = tk.StringVar(value="Efectivo")
        ttk.Radiobutton(frame_pago, text="Efectivo", value="Efectivo", variable=self.forma_pago).pack(side="left", padx=5)
        ttk.Radiobutton(frame_pago, text="Tarjeta", value="Tarjeta", variable=self.forma_pago).pack(side="left", padx=5)
        ttk.Radiobutton(frame_pago, text="Transferencia", value="Transferencia", variable=self.forma_pago).pack(side="left", padx=5)
        # -------------------------
        # Botón guardar Excel y actualizar inventario
        # -------------------------
        ttk.Button(self, text="Guardar Venta y Actualizar Inventario", command=self.guardar_excel_y_actualizar).pack(pady=10)
        # ------------------
        # Tabla de venta con columnas
        # ------------------
        cols = ["forma_pago", "codigo", "cantidad", "p_unitario", "precio", "total"]
        self.tree = ttk.Treeview(self, columns=cols, show='headings', height=14)
        for c in cols:
            self.tree.heading(c, text=c.replace("_", " ").capitalize())
            if c in ["codigo"]:
                self.tree.column(c, width=100, anchor='center')
            else:
                self.tree.column(c, width=120, anchor='center')
        self.tree.pack(fill='both', expand=True, padx=6, pady=6)
        # ------------------
        # Botón borrar seleccionado
        # ------------------
        ttk.Button(self, text="Borrar Seleccionado", command=self.borrar_seleccionado).pack(pady=5)
    # -----------------
    # Función para autocompletar desde inventario.xlsx
    # ------------------
    def completar_datos(self, event=None):
        codigo = self.cod_entry.get().strip().upper()
        if not codigo:
            return
        df = load_inventario_file()  # Lee inventario.xlsx o lo crea vacío si no existe
        mask = df['codigo'].astype(str).str.upper() == codigo
        if mask.any():
            fila = df[mask].iloc[0]
            self.desc_entry.delete(0, tk.END)
            self.desc_entry.insert(0, fila['descripcion'])
            self.precio_entry.delete(0, tk.END)
            self.precio_entry.insert(0, fila['precio'])
        else:
            self.desc_entry.delete(0, tk.END)
            self.precio_entry.delete(0, tk.END)
    # ------------------
    # Agregar producto a la tabla
    # ------------------
    def agregar_a_venta(self):
        codigo = self.cod_entry.get().strip().upper()
        try:
            cantidad = int(self.cant_entry.get())
        except:
            cantidad = 0
        try:
            precio = float(self.precio_entry.get())
        except:
            precio = 0.0
        if not codigo or cantidad <= 0 or precio <= 0:
            messagebox.showwarning("Atención", "Código, cantidad y precio deben ser válidos")
            return
        p_unitario = round(precio / 1.16, 2)  # Precio unitario sin IVA
        total = cantidad * precio
        forma = self.forma_pago.get()
        self.tree.insert('', 'end', values=(forma, codigo, cantidad, p_unitario, precio, total))
        # Limpiar campos
        self.cod_entry.delete(0, tk.END)
        self.desc_entry.delete(0, tk.END)
        self.precio_entry.delete(0, tk.END)
        self.cant_entry.delete(0, tk.END)
    # -----------------
    # Borrar seleccionado
    # ------------------
    def borrar_seleccionado(self):
        seleccionado = self.tree.selection()
        if not seleccionado:
            messagebox.showwarning("Atención", "Seleccione un producto para borrar.")
            return
        for item in seleccionado:
            self.tree.delete(item)
    # ------------------
    # Guardar en Excel y actualizar inventario
    # ------------------
    def guardar_excel_y_actualizar(self):
        try:
            productos = [self.tree.item(i)["values"] for i in self.tree.get_children()]
            if not productos:
                messagebox.showwarning("Atención", "No hay productos en la venta.")
                return
            # Crear DataFrame con las columnas correctas
            df_nuevo = pd.DataFrame(productos, columns=["Forma_Pago", "Código", "Cantidad", "P_Unitario", "Precio", "Total"])
            # Leer archivo existente o crear uno nuevo
            if ARCHIVO_VENTAS.exists():
                df_exist = pd.read_excel(ARCHIVO_VENTAS, engine="openpyxl")
                df_comb = pd.concat([df_exist, df_nuevo], ignore_index=True)
            else:
                df_comb = df_nuevo
            # Ordenar por forma de pago
            orden = ["Efectivo", "Tarjeta", "Transferencia"]
            df_comb["Forma_Pago"] = pd.Categorical(df_comb["Forma_Pago"], categories=orden, ordered=True)
            df_comb = df_comb.sort_values("Forma_Pago")
            # Guardar ventas
            df_comb.to_excel(ARCHIVO_VENTAS, index=False, engine="openpyxl")
            # Actualizar inventario
            df_inv = load_inventario_file()
            for _, row in df_nuevo.iterrows():
                mask = df_inv['codigo'].astype(str).str.upper() == row['Código'].upper()
                if mask.any():
                    idx = df_inv[mask].index[0]
                    df_inv.at[idx, 'stock'] = max(0, int(df_inv.at[idx, 'stock']) - int(row['Cantidad']))
            save_inventario_file(df_inv)
            messagebox.showinfo("Éxito", f"Venta guardada y stock actualizado.\nArchivo: {ARCHIVO_VENTAS}")
            # Limpiar tabla
            for i in self.tree.get_children():
                self.tree.delete(i)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar o actualizar inventario:\n{e}")
# ===============================
# COTIZACION
# ===============================
class Cotizacion(ttk.Frame):
    def __init__(self, parent, controller=None, inventario_df=None):
        super().__init__(parent)
        self.controller = controller
        self.inventario_df = inventario_df
        # ------------------------
        # Variables
        # ------------------------
        self.total_parcial_var = tk.StringVar(value="0.00")
        self.total_general_var = tk.StringVar(value="0.00")
        # ------------------------
        # Título
        # ------------------------
        ttk.Label(self, text='COTIZACIÓN', font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=6, pady=6)
        # ------------------------
        # Frame de búsqueda / agregar
        # ------------------------
        frame_buscar = ttk.LabelFrame(self, text="Agregar Producto a Cotización", padding=10)
        frame_buscar.pack(fill="x", padx=10, pady=10)
        # ---- Código ----
        ttk.Label(frame_buscar, text="Código:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.entry_codigo = ttk.Entry(frame_buscar, width=25)
        self.entry_codigo.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.entry_codigo.bind("<KeyRelease>", self.autocompletar_producto)
        # ---- Descripción ----
        ttk.Label(frame_buscar, text="Descripción:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.entry_desc = ttk.Entry(frame_buscar, width=40)
        self.entry_desc.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        # ---- Precio ----
        ttk.Label(frame_buscar, text="Precio:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.entry_precio = ttk.Entry(frame_buscar, width=20)
        self.entry_precio.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.entry_precio.bind("<KeyRelease>", self.actualizar_total_parcial)
        # ---- Stock ----
        ttk.Label(frame_buscar, text="Cantidad disponible:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.entry_stock = ttk.Entry(frame_buscar, width=20)
        self.entry_stock.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        # ---- Cantidad ----
        ttk.Label(frame_buscar, text="Cantidad a cotizar:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.entry_cantidad = ttk.Entry(frame_buscar, width=10)
        self.entry_cantidad.grid(row=4, column=1, padx=5, pady=5, sticky="w")
        self.entry_cantidad.bind("<KeyRelease>", self.actualizar_total_parcial)
        # ---- Disponibilidad ----
        ttk.Label(frame_buscar, text="Disponibilidad:").grid(row=5, column=0, padx=5, pady=5, sticky="e")
        self.combo_disp = ttk.Combobox(frame_buscar, values=["Disponible", "No disponible"], width=18)
        self.combo_disp.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        self.combo_disp.set("Disponible")
        # ---- Total parcial dinámico ----
        ttk.Label(frame_buscar, text="TOTAL: ", font=("Arial", 10, "bold")).grid(row=6, column=0, pady=5, sticky="e")
        ttk.Label(frame_buscar, textvariable=self.total_parcial_var, font=("Arial", 12, "bold"), foreground="green").grid(row=6, column=1, pady=5, sticky="w")
        # ---- Botón agregar ----
        ttk.Button(frame_buscar, text="Agregar a Cotización", command=self.agregar_producto).grid(
            row=7, column=0, columnspan=2, pady=10
        )
        # ------------------------
        # Treeview
        # ------------------------
        columnas = ("Código", "Descripción", "Precio", "Cantidad", "Total", "Disponibilidad")
        self.tree = ttk.Treeview(self, columns=columnas, show="headings", height=12)
        for col in columnas:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        habilitar_copia_treeview(self.tree)
        # ------------------------
        # Total general
        # ------------------------
        frame_tg = ttk.Frame(self)
        frame_tg.pack(fill="x", padx=10, pady=5)
        ttk.Label(frame_tg, text="Total general:", font=("Arial", 10, "bold")).pack(side="left")
        ttk.Label(frame_tg, textvariable=self.total_general_var, font=("Arial", 10, "bold")).pack(side="left", padx=8)
        # ------------------------
        # Botones finales
        # ------------------------
        frame_acciones_final = ttk.Frame(self)
        frame_acciones_final.pack(fill="x", padx=10, pady=10)
        ttk.Button(frame_acciones_final, text="Eliminar Seleccionado", command=self.eliminar_producto).pack(side="left", padx=5)
        ttk.Button(frame_acciones_final, text="Guardar en Excel", command=self.guardar_excel).pack(side="left", padx=5)
        ttk.Button(frame_acciones_final, text="Crear Ticket PDF", command=self.crear_ticket_pdf).pack(side="left", padx=5)
    # =====================================================
    # AUTOCOMPLETAR PRODUCTO
    # =====================================================
    def autocompletar_producto(self, event=None):
        codigo = quitar_acentos(self.entry_codigo.get().strip()).upper()
        if len(codigo) < 1:
            return
        try:
            df = load_inventario_file()
            df.columns = df.columns.str.strip().str.lower()
            df.fillna("", inplace=True)
            if "codigo" not in df.columns:
                df["codigo"] = df.iloc[:, 0]
            df["codigo_clean"] = df["codigo"].apply(lambda x: quitar_acentos(str(x)).upper())
            resultado = df[df["codigo_clean"] == codigo]
            if resultado.empty:
                resultado = df[df["codigo_clean"].str.contains(codigo, na=False)]
                if resultado.empty:
                    for e in [self.entry_desc, self.entry_precio, self.entry_stock]:
                        e.delete(0, tk.END)
                    self.total_parcial_var.set("0.00")
                    return
            fila = resultado.iloc[0]
            self.entry_desc.delete(0, tk.END)
            self.entry_desc.insert(0, fila.get("descripcion", ""))
            self.entry_precio.delete(0, tk.END)
            self.entry_precio.insert(0, str(fila.get("precio", "")))
            self.entry_stock.delete(0, tk.END)
            self.entry_stock.insert(0, str(fila.get("stock", "")))
            self.actualizar_total_parcial()
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar inventario: {e}")
    # =====================================================
    # TOTAL PARCIAL AUTOMÁTICO
    # =====================================================
    def actualizar_total_parcial(self, event=None):
        try:
            precio = float(self.entry_precio.get())
        except:
            precio = 0
        try:
            cantidad = float(self.entry_cantidad.get())
        except:
            cantidad = 0
        total = precio * cantidad
        self.total_parcial_var.set(f"{total:,.2f}")
    # =====================================================
    # AGREGAR PRODUCTO A COTIZACIÓN
    # =====================================================
    def agregar_producto(self):
        try:
            codigo = self.entry_codigo.get().strip()
            desc = self.entry_desc.get().strip()
            precio = float(self.entry_precio.get())
            cantidad = int(self.entry_cantidad.get())
            total = precio * cantidad
            disp = self.combo_disp.get().strip()
            self.tree.insert("", tk.END, values=(codigo, desc, f"{precio:.2f}", cantidad, f"{total:.2f}", disp))
            for e in [self.entry_codigo, self.entry_desc, self.entry_precio, self.entry_stock, self.entry_cantidad]:
                e.delete(0, tk.END)
            self.combo_disp.set("Disponible")
            self.total_parcial_var.set("0.00")
            self.recalcular_total_general()
        except Exception as e:
            messagebox.showwarning("Atención", f"Error al agregar producto: {e}")
    # =====================================================
    # ELIMINAR PRODUCTO
    # =====================================================
    def eliminar_producto(self):
        sel = self.tree.selection()
        for item in sel:
            self.tree.delete(item)
        self.recalcular_total_general()
    # =====================================================
    # TOTAL GENERAL
    # =====================================================
    def recalcular_total_general(self):
        total = 0.0
        for item in self.tree.get_children():
            try:
                total += float(self.tree.item(item)["values"][4])
            except:
                pass
        self.total_general_var.set(f"{total:,.2f}")
    # =====================================================
    # GUARDAR EXCEL
    # =====================================================
    def guardar_excel(self):
        try:
            productos = [self.tree.item(i)["values"] for i in self.tree.get_children()]
            if not productos:
                messagebox.showwarning("Atención", "No hay productos para guardar.")
                return
            df = pd.DataFrame(productos, columns=["Código", "Descripción", "Precio", "Cantidad", "Total", "Disponibilidad"])
            archivo = "cotizacion.xlsx"
            df.to_excel(archivo, index=False)
            messagebox.showinfo("Éxito", f"Cotización guardada en {archivo}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar en Excel: {e}")
    # =====================================================
    # CREAR TICKET PDF
    # =====================================================
    def crear_ticket_pdf(self):
        try:
            productos = [self.tree.item(i)["values"] for i in self.tree.get_children()]
            if not productos:
                messagebox.showwarning("Atención", "No hay productos para exportar.")
                return
            from reportlab.pdfgen import canvas
            archivo = "ticket_cotizacion.pdf"
            c = canvas.Canvas(archivo)
            y = 800
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, y, "COTIZACIÓN")
            y -= 30
            c.setFont("Helvetica", 10)
            for p in productos:
                linea = f"{p[0]} | {p[1]} | {p[3]} x ${p[2]} = ${p[4]}"
                c.drawString(40, y, linea)
                y -= 20
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y-10, f"TOTAL GENERAL: ${self.total_general_var.get()}")
            c.save()
            messagebox.showinfo("Éxito", f"PDF creado: {archivo}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el PDF: {e}")
class Taller(ttk.Frame):
    def __init__(self, parent, controller=None):
        super().__init__(parent)
        self.controller = controller
        self.motos = {}  # {"Moto1": [{"codigo":.., "descripcion":.., "precio":.., "cantidad":.., "total":..}, ...]}
        tk.Label(self, text="TALLER", font=("Arial", 20), bg="white").pack(pady=10)
        frame_top = ttk.Frame(self)
        frame_top.pack(fill="x", padx=10, pady=(20,5))
        frame_botones = ttk.Frame(frame_top)
        frame_botones.pack(side="left", anchor="w")       
        ttk.Button(frame_botones, text="📥 Importar Archivo", command=self.importar_archivo).pack(side="top", pady=2)
        ttk.Button(frame_botones, text="Exportar Excel", command=self.exportar_excel).pack(side="top", pady=2)
        ttk.Button(frame_botones, text="Crear PDF", command=self.crear_pdf).pack(side="top", pady=2)
        # Treeview de motos
        self.tree_motos = ttk.Treeview(frame_top, columns=("Moto", "Total"), show="headings", height=6)
        self.tree_motos.heading("Moto", text="Moto")
        self.tree_motos.heading("Total", text="Total")
        self.tree_motos.column("Moto", width=200)
        self.tree_motos.column("Total", width=100, anchor="center")
        self.tree_motos.pack(side="left", fill="x", expand=True, padx=5)
        # Acciones
        frame_acciones = ttk.Frame(frame_top)
        frame_acciones.pack(side="right", padx=5)
        ttk.Button(frame_acciones, text="Agregar Moto", command=self.agregar_moto).pack(fill="x", pady=2)
        ttk.Button(frame_acciones, text="Borrar Moto", command=self.borrar_moto).pack(fill="x", pady=2)
        ttk.Button(frame_acciones, text="Agregar Insumo", command=self.agregar_insumo).pack(fill="x", pady=2)
        ttk.Button(frame_acciones, text="Guardar Taller", command=self.guardar_taller).pack(fill="x", pady=2)       
        self.cargar_taller()
    # -------------------------- MÉTODOS DE MOTO --------------------------
    def agregar_moto(self):
        nombre = simpledialog.askstring("Nueva Moto", "Nombre de la moto:")
        if not nombre:
            return
        nombre = nombre.strip()
        if nombre in self.motos:
            messagebox.showwarning("Atención", "La moto ya existe.")
            return
        self.motos[nombre] = []
        self.tree_motos.insert("", "end", iid=nombre, values=(nombre, "0.00"))
    def borrar_moto(self):
        sel = self.tree_motos.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccione una moto para borrar.")
            return
        for moto in sel:
            self.motos.pop(moto, None)
            self.tree_motos.delete(moto)

    # -------------------------- AGREGAR INSUMO --------------------------
    def agregar_insumo(self):
        sel = self.tree_motos.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccione primero una moto.")
            return
        moto = sel[0]
        win = tk.Toplevel(self)
        win.title(f"Insumos - {moto}")
        win.geometry("800x650")
        # Variables
        codigo_var = tk.StringVar()
        descripcion_var = tk.StringVar()
        precio_var = tk.StringVar()
        cantidad_var = tk.StringVar(value="1")
        total_var = tk.StringVar(value="0.00")
        ttk.Label(win, text="Código").grid(row=0, column=0, padx=6, pady=6)
        entry_codigo = ttk.Entry(win, textvariable=codigo_var)
        entry_codigo.grid(row=0, column=1, padx=6, pady=6)
        ttk.Label(win, text="Cantidad").grid(row=1, column=0, padx=6, pady=6)
        entry_cantidad = ttk.Entry(win, textvariable=cantidad_var)
        entry_cantidad.grid(row=1, column=1, padx=6, pady=6)
        ttk.Label(win, text="Descripción").grid(row=2, column=0, padx=6, pady=6)
        entry_desc = ttk.Entry(win, textvariable=descripcion_var, width=40)
        entry_desc.grid(row=2, column=1, columnspan=3, padx=6, pady=6)
        ttk.Label(win, text="Precio").grid(row=3, column=0, padx=6, pady=6)
        entry_precio = ttk.Entry(win, textvariable=precio_var)
        entry_precio.grid(row=3, column=1, padx=6, pady=6)
        ttk.Label(win, text="Total").grid(row=4, column=0, padx=6, pady=6)
        entry_total = ttk.Entry(win, textvariable=total_var, state="readonly")
        entry_total.grid(row=4, column=1, padx=6, pady=6)
        # Treeview de insumos
        cols = ("codigo", "cantidad", "descripcion", "precio", "total")
        tree_ins = ttk.Treeview(win, columns=cols, show="headings", height=10)
        for c in cols:
                tree_ins.heading(c, text=c.capitalize())
                tree_ins.column(c, width=100, anchor="center")
        tree_ins.grid(row=5, column=0, columnspan=4, padx=6, pady=6, sticky="nsew")

        # Función de autocompletar descripción y precio
        def actualizar_autocompletar(*args):
                codigo = codigo_var.get().strip().upper()
                if not codigo:
                        descripcion_var.set("")
                        precio_var.set("0.00")
                        return
                try:
                        df = pd.read_excel(ARCHIVO_INVENTARIO, engine="openpyxl")
                        df['codigo'] = df['codigo'].astype(str)
                        prod = df[df['codigo'].str.upper() == codigo]
                        if not prod.empty:
                                fila = prod.iloc[0]
                                descripcion_var.set(fila["descripcion"])
                                precio_var.set(str(fila["precio"]))
                        else:
                                descripcion_var.set("")
                                precio_var.set("0.00")
                except:
                        descripcion_var.set("")
                        precio_var.set("0.00")
                actualizar_total()
        codigo_var.trace("w", actualizar_autocompletar)
        # Función actualizar total
        def actualizar_total(*args):
                try:
                        total = float(precio_var.get()) * float(cantidad_var.get())
                        total_var.set(f"{total:.2f}")
                except:
                        total_var.set("0.00")
        cantidad_var.trace("w", actualizar_total)
        precio_var.trace("w", actualizar_total)
        # Función agregar insumo al tree y al dict
        def agregar_local():
                try:
                        cant = float(cantidad_var.get())
                        precio = float(precio_var.get())
                        total = float(total_var.get())
                        codigo = codigo_var.get().strip().upper()
                        desc = descripcion_var.get()
                except:
                        messagebox.showwarning("Atención", "Cantidad o precio inválidos")
                        return
                # Revisar si ya existe en Treeview
                for iid in tree_ins.get_children():
                        vals = tree_ins.item(iid)["values"]
                        if vals[0] == codigo:
                                nueva_cant = float(vals[1]) + cant
                                nuevo_total = round(nueva_cant * precio, 2)
                                tree_ins.item(iid, values=(codigo, nueva_cant, desc, precio, nuevo_total))
                                break
                else:
                        tree_ins.insert("", "end", values=(codigo, cant, desc, precio, total))
                # Actualizar dict
                for ins in self.motos[moto]:
                        if ins["codigo"] == codigo:
                                ins["cantidad"] += cant
                                ins["precio"] = precio
                                ins["total"] = round(ins["cantidad"] * precio, 2)
                                break
                else:
                        self.motos[moto].append({
                                "codigo": codigo,
                                "cantidad": cant,
                                "descripcion": desc,
                                "precio": precio,
                                "total": total
                        })
                # Actualizar total de moto
                total_moto = sum(i["total"] for i in self.motos[moto])
                self.tree_motos.item(moto, values=(moto, f"{total_moto:.2f}"))
                # Limpiar entradas
                codigo_var.set("")
                descripcion_var.set("")
                precio_var.set("0.00")
                cantidad_var.set("1")
                total_var.set("0.00")
        # Función borrar insumo seleccionado
        def borrar_seleccionado():
                seleccionado = tree_ins.selection()
                if not seleccionado:
                        messagebox.showwarning("Atención", "Seleccione un insumo para borrar.")
                        return
                for iid in seleccionado:
                        vals = tree_ins.item(iid)["values"]
                        codigo_borrar = vals[0]
                        # Quitar del Treeview
                        tree_ins.delete(iid)
                        # Quitar del dict
                        self.motos[moto] = [ins for ins in self.motos[moto] if ins["codigo"] != codigo_borrar]
                # Actualizar total de la moto
                total_moto = sum(i["total"] for i in self.motos[moto])
                self.tree_motos.item(moto, values=(moto, f"{total_moto:.2f}"))
        # Botones
        ttk.Button(win, text="Agregar", command=agregar_local).grid(row=6, column=0, pady=6, padx=6)
        ttk.Button(win, text="Borrar Seleccionado", command=borrar_seleccionado).grid(row=6, column=1, pady=6, padx=6)
        ttk.Button(win, text="Cerrar", command=win.destroy).grid(row=6, column=2, pady=6, padx=6)
    def guardar_motos(self):
        from openpyxl import load_workbook
        import pandas as pd
        if ARCHIVO_MOTOS.exists():
                wb = load_workbook(ARCHIVO_MOTOS)
        else:
                wb = None
        with pd.ExcelWriter(ARCHIVO_MOTOS, engine="openpyxl") as writer:
                if wb:
                        writer.book = wb
                        # Remover sheets existentes de motos para reemplazar
                        for moto in self.motos.keys():
                                if moto in writer.book.sheetnames:
                                        idx = writer.book.sheetnames.index(moto)
                                        ws = writer.book.worksheets[idx]
                                        writer.book.remove(ws)
                for moto, insumos in self.motos.items():
                        df = pd.DataFrame(insumos, columns=["codigo","descripcion","cantidad","precio","total"])
                        df.to_excel(writer, sheet_name=moto, index=False)
    # -------------------------- GUARDAR TALLER --------------------------
    def guardar_taller(self):
        if not self.motos:
            messagebox.showwarning("Atención", "No hay motos para guardar.")
            return
        os.makedirs(os.path.dirname(ARCHIVO_TALLER), exist_ok=True)
        writer = pd.ExcelWriter(ARCHIVO_TALLER, engine="openpyxl")
        for moto, insumos in self.motos.items():
            if insumos:
                df = pd.DataFrame(insumos)
                df.to_excel(writer, sheet_name=moto[:31], index=False)
        writer.close()
        messagebox.showinfo("Éxito", "Taller guardado correctamente.")
    # -------------------------- CARGAR TALLER --------------------------
    def cargar_taller(self):
        if not os.path.exists(ARCHIVO_TALLER):
            return
        xls = pd.ExcelFile(ARCHIVO_TALLER, engine="openpyxl")
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")
            # Crear columnas si no existen
            if "cantidad" not in df.columns:
                df["cantidad"] = 0
            if "precio" not in df.columns:
                df["precio"] = 0
            # Convertir a numérico y llenar NaN con 0
            df["cantidad"] = pd.to_numeric(df["cantidad"], errors="coerce").fillna(0)
            df["precio"] = pd.to_numeric(df["precio"], errors="coerce").fillna(0)
            df["total"] = df["cantidad"] * df["precio"]
            self.motos[sheet] = df.to_dict("records")
            total_moto = df["total"].sum()
            self.tree_motos.insert("", "end", iid=sheet, values=(sheet, f"{total_moto:.2f}"))
    # -------------------------- EXPORTAR EXCEL --------------------------
    def exportar_excel(self):
        sel = self.tree_motos.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccione una moto")
            return
        moto = sel[0]
        insumos = self.motos[moto]
        archivo = f"{moto}_taller.xlsx"
        pd.DataFrame(insumos).to_excel(archivo, index=False)
        messagebox.showinfo("Éxito", f"Excel exportado: {archivo}")
    # -------------------------- CREAR PDF --------------------------
    def crear_pdf(self):
        sel = self.tree_motos.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccione una moto")
            return
        moto = sel[0]
        insumos = self.motos[moto]
        archivo = f"{moto}_taller.pdf"
        c = canvas.Canvas(archivo, pagesize=letter)
        y = 750
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, f"Taller - {moto}")
        y -= 30
        total_general = 0
        for i in insumos:
            c.drawString(40, y, f'{i["codigo"]} - {i["descripcion"]} - {i["cantidad"]} x {i["precio"]} = {i["total"]}')
            total_general += i["total"]
            y -= 20
        c.drawString(40, y, f"Total: {total_general:.2f}")
        c.save()
        messagebox.showinfo("Éxito", f"PDF creado: {archivo}")
    # -----------------
    # Importar Excel
    # -----------------
    def importar_archivo(self):
        ruta = filedialog.askopenfilename(title="Seleccionar archivo Excel",
                                          filetypes=[("Archivos Excel", "*.xlsx *.xls")])
        if not ruta:
            return
        try:
            xls = pd.ExcelFile(ruta, engine="openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:{str(e)}")
            return
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")
                df["cantidad"] = pd.to_numeric(df.get("cantidad", 0), errors="coerce").fillna(0)
                df["precio"] = pd.to_numeric(df.get("precio", 0), errors="coerce").fillna(0)
                df["total"] = df["cantidad"] * df["precio"]
                self.motos[sheet] = df.to_dict("records")
                total_moto = df["total"].sum()
                if not self.tree_motos.exists(sheet):
                    self.tree_motos.insert("", "end", iid=sheet, values=(sheet, f"{total_moto:.2f}"))
                else:
                    self.tree_motos.item(sheet, values=(sheet, f"{total_moto:.2f}"))
            except Exception as e:
                messagebox.showwarning("Atención", f"No se pudo cargar la hoja '{sheet}': {e}")
        messagebox.showinfo("Éxito", "Archivo importado correctamente.")
# --------------------
# MAIN APP
# --------------------
import threading
from flask import Flask, jsonify

# --------------------
class AppUnificada(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("JQ MOTORS SISTEM")
        self.geometry("1300x800")
        self.configure(bg="white")
        # DataFrame de inventario (inicializar vacío o cargar desde Excel/CSV)
        self.inventario_df = pd.DataFrame()  # reemplaza con tu carga real si tienes archivo
        # Definir la variable del total
        self.total_var = tk.StringVar()
        self.total_var.set("0.00")
        # HEADER
        header = ttk.Frame(self)
        header.pack(fill='x', pady=(5,0))
        ttk.Label(header, text='🚀 JQ MOTORS MATRIZ', font=('Segoe UI',14,'bold')).pack(side='left', padx=10)
        # Notebook principal
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True)
        # Crear instancias de cada pestaña
        self.stock_tab = Stock(self.notebook, controller=self)
        self.ventas_tab = Ventas(self.notebook)
        self.cot_tab = Cotizacion(self.notebook, controller=self, inventario_df=self.inventario_df)
        self.taller_tab = Taller(self.notebook)
        # Añadir pestañas al notebook
        self.notebook.add(self.stock_tab, text='Stock')
        self.notebook.add(self.ventas_tab, text='Ventas')
        self.notebook.add(self.cot_tab, text='Cotizacion')
        self.notebook.add(self.taller_tab, text='Taller')
        # Frame principal para entradas
        self.frame = tk.Frame(self, bg="white")
        self.frame.pack(fill="both", expand=True, padx=10, pady=10)
        # Cambiar estilo de Treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="white", foreground="#003366", fieldbackground="white", rowheight=25)
        style.configure("Treeview.Heading", background="#003366", foreground="white", font=("Arial", 10, "bold"))
        # Lista interna de registros
        self.registros = []

        # --------------------
        # Establecer referencia global para Flask
        # --------------------
        global APP_GLOBAL
        APP_GLOBAL = self

# --------------------
# SERVIDOR FLASK
# --------------------
app_flask = Flask(__name__)

@app_flask.route('/inventario', methods=['GET'])
def inventario_json():
    try:
        if 'APP_GLOBAL' in globals():
            df = APP_GLOBAL.inventario_df
            data = df.to_dict(orient='records')
            return jsonify(data)
        else:
            return jsonify([])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def iniciar_servidor_flask():
    app_flask.run(host='0.0.0.0', port=5002, debug=False)

# --------------------
# MAIN
# --------------------
if __name__ == "__main__":
    # Iniciar servidor Flask en segundo plano
    flask_thread = threading.Thread(target=iniciar_servidor_flask, daemon=True)
    flask_thread.start()

    # Iniciar app Tkinter
    app = AppUnificada()
    app.mainloop()
